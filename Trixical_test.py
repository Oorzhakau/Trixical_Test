from matplotlib import pyplot as plt
from scipy.interpolate import interp1d
import numpy as np
import shutil
import openpyxl

#book = xlwt.Workbook(encoding="utf-8")
#sheet1 = book.add_sheet("Sheet 1")

n = 100 

class LineBuilder:
    def __init__(self, line, bh = 'x', depth = '1', sigma3 = 0.1):
        self.line = line
        self.bh = bh
        self.depth = depth
        self.sigma3 = sigma3
        self.xs = list(line.get_xdata())
        self.ys = list(line.get_ydata())
        self.cid = line.figure.canvas.mpl_connect('button_press_event', self)
        self.cid1 = line.figure.canvas.mpl_connect('key_press_event', self.press)
        
    def __call__(self, event):
        print('click', event)
        if event.inaxes!=self.line.axes: return
        self.xs.append(event.xdata)
        self.ys.append(event.ydata)
        self.line.set_data(self.xs, self.ys)
        self.line.figure.canvas.draw()
        
    def press(self, event):
        if event.key == 'x':
            x1 = self.xs[0]
            y1 = self.ys[0]
            self.xs = list()
            self.ys = list()
            self.xs.append(x1)
            self.ys.append(y1)
            self.line.set_data(self.xs, self.ys)
            self.line.figure.canvas.draw()
        if event.key == 'w':
            f = interp1d(self.xs, self.ys, kind='cubic')
            xnew = np.linspace(self.xs[0], self.xs[-1], num = n)
            ynew = f(xnew)
            self.line.set_data(xnew, ynew)
            self.line.figure.canvas.draw()
            shutil.copy('Exp.xlsx',self.bh+"_"+self.depth+".xlsx")
            self.wb = openpyxl.load_workbook(self.bh+"_"+self.depth+".xlsx")
            
            self.ws = self.wb['Sheet 1']
            self.ws['C20'] = self.bh 
            self.ws['C21'] = self.depth
            self.ws['C27'] = self.sigma3
            for num in range(len(xnew)):
                self.ws['R'+str(10+num)] = xnew[num]
                self.ws['S'+str(10+num)] = ynew[num]
                #self.ws.cell(row = (9 + num), column = 18, value = xnew[num])
                #self.ws.cell(row = (9 + num), column = 19, value = ynew[num])
            self.wb.save(self.bh+"_"+self.depth+".xlsx")

if __name__ == "__main__":
    fig = plt.figure()
    ax = fig.add_subplot(111)
    ax.set_title('Click on canvas to paint line\nPress "x" to clear canvas or "w" to interpolate and save data')

    line, = ax.plot([], [])  # empty line
    linebuilder = LineBuilder(line)
    xl = ax.set_xlabel('')
    plt.xlim(0, 0.16)
    plt.ylim(0, 0.6)
    plt.xlabel('eps, d.e.')
    plt.ylabel('Sigma1 - Sigma3, MPa')
    plt.show()